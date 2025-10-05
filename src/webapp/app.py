from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import mysql.connector
from mysql.connector import Error
import os
from datetime import datetime
import logging
import uuid
import zipfile
import io
import pandas as pd
from openpyxl import load_workbook

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Configuración
DB_CONFIG = {
    'host': 'localhost',
    'database': 'nasa',
    'user': 'root',
    'password': ''
}

# Configuración de carpetas
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'txt', 'csv', 'json', 'xlsx'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Crear carpeta de uploads si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db_connection():
    """Crear conexión a la base de datos"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        logger.error(f"Error conectando a MySQL: {e}")
        return None

def allowed_file(filename):
    """Verificar si la extensión del archivo es permitida"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_file_to_disk(file):
    """Guardar archivo en el sistema de archivos y retornar la ruta"""
    try:
        # Generar nombre único para evitar colisiones
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        
        # Guardar archivo
        file.save(file_path)
        return file_path, unique_filename
        
    except Exception as e:
        logger.error(f"Error guardando archivo: {e}")
        return None, None
    
def procesar_excel_plantilla(file_path):
    """Procesar archivo Excel según la plantilla de exoplanetas"""
    try:
        # Cargar el archivo Excel
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Mapeo de filas a nombres de columna_final (A2:A12)
        mapeo_columnas = {
            2: "type",
            3: "ra", 
            4: "dec",
            5: "pl_orbper",
            6: "pl_rade",
            7: "pl_insol",
            8: "pl_eqt",
            9: "st_teff",
            10: "st_logg",
            11: "st_rad",
            12: "st_tmag"
        }
        
        datos_exoplanetas = []
        
        for row in range(2, 13):  # Filas 2 a 12
            # Obtener el nombre de la columna_final de la celda A
            columna_final = sheet[f'A{row}'].value
            
            # Si la celda A está vacía, usar el mapeo por defecto
            if columna_final is None or columna_final == '':
                columna_final = mapeo_columnas.get(row, f"fila_{row}")
            else:
                columna_final = str(columna_final).strip()
            
            fila_datos = {
                'columna_final': columna_final,
                'fila_excel': row  # Mantenemos esto como referencia
            }
            
            # Columna B: Origen TESS
            valor_b = sheet[f'B{row}'].value
            if valor_b is not None and valor_b != '—' and valor_b != '':
                fila_datos['origen_tess'] = str(valor_b).strip()
            
            # Columna C: Origen Kepler
            valor_c = sheet[f'C{row}'].value
            if valor_c is not None and valor_c != '—' and valor_c != '':
                fila_datos['origen_kepler'] = str(valor_c).strip()
            
            # Columna D: Origen K2
            valor_d = sheet[f'D{row}'].value
            if valor_d is not None and valor_d != '—' and valor_d != '':
                fila_datos['origen_k2'] = str(valor_d).strip()
            
            # Columna E: Descripción
            valor_e = sheet[f'E{row}'].value
            if valor_e is not None and valor_e != '':
                fila_datos['descripcion'] = str(valor_e).strip()
            
            # Solo agregar filas que tengan datos en al menos una columna B, C, D o E
            tiene_datos = any(key in fila_datos for key in ['origen_tess', 'origen_kepler', 'origen_k2', 'descripcion'])
            
            if tiene_datos:
                datos_exoplanetas.append(fila_datos)
        
        return datos_exoplanetas
        
    except Exception as e:
        logger.error(f"Error procesando Excel: {e}")
        raise Exception(f"Error al procesar archivo Excel: {str(e)}")

def extraer_metadatos_excel(file_path):
    """Extraer metadatos adicionales del archivo Excel"""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        metadatos = {
            'total_filas_procesadas': 0,
            'columnas_detectadas': [],
            'columnas_finales': [],
            'primera_fila_con_datos': None,
            'ultima_fila_con_datos': None
        }
        
        # Contar filas con datos en el rango B2:E12 y recolectar nombres de columnas finales
        filas_con_datos = 0
        columnas_finales = []
        
        for row in range(2, 13):
            tiene_datos = any(
                sheet[col + str(row)].value is not None and 
                sheet[col + str(row)].value != '—' and 
                sheet[col + str(row)].value != ''
                for col in ['B', 'C', 'D', 'E']
            )
            
            if tiene_datos:
                filas_con_datos += 1
                # Obtener nombre de columna final
                columna_final = sheet[f'A{row}'].value
                if columna_final:
                    columnas_finales.append(str(columna_final).strip())
                
                if metadatos['primera_fila_con_datos'] is None:
                    metadatos['primera_fila_con_datos'] = row
                metadatos['ultima_fila_con_datos'] = row
        
        metadatos['total_filas_procesadas'] = filas_con_datos
        metadatos['columnas_detectadas'] = ['origen_tess', 'origen_kepler', 'origen_k2', 'descripcion']
        metadatos['columnas_finales'] = columnas_finales
        
        return metadatos
        
    except Exception as e:
        logger.error(f"Error extrayendo metadatos Excel: {e}")
        return {}

@app.route('/upload', methods=['POST'])
def upload_files():
    """Endpoint para subir archivos y procesar datos de Excel"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No se encontraron archivos'}), 400
        
        files = request.files.getlist('files')
        donador = request.form.get('donador', 'Anónimo')
        descripcion = request.form.get('description', '')
        consentimiento = request.form.get('consent') == 'true'
        
        if not files or files[0].filename == '':
            return jsonify({'error': 'No se seleccionaron archivos'}), 400
        
        if not consentimiento:
            return jsonify({'error': 'Debe aceptar los términos de uso'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor()
        archivos_subidos = []
        archivos_con_error = []
        datos_procesados = []  # Para almacenar datos extraídos
        
        for file in files:
            # Validar tipo de archivo
            if not allowed_file(file.filename):
                archivos_con_error.append({
                    'nombre': file.filename,
                    'error': 'Tipo de archivo no permitido'
                })
                continue
            
            # Validar tamaño
            file.seek(0, 2)  # Ir al final
            file_size = file.tell()
            file.seek(0)  # Volver al inicio
            
            if file_size > MAX_FILE_SIZE:
                archivos_con_error.append({
                    'nombre': file.filename,
                    'error': 'Archivo demasiado grande'
                })
                continue
            
            # Guardar archivo en carpeta
            file_path, unique_filename = save_file_to_disk(file)
            if not file_path:
                archivos_con_error.append({
                    'nombre': file.filename,
                    'error': 'Error guardando archivo'
                })
                continue
            
            # Insertar metadata en la base de datos
            sql_documento = """
            INSERT INTO documentos_exoplanetas 
            (nombre_archivo, tipo_archivo, tamano_archivo, ruta_archivo, donador, descripcion, consentimiento)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            file_extension = file.filename.rsplit('.', 1)[1].lower()
            
            cursor.execute(sql_documento, (
                file.filename,          # Nombre original
                file_extension,         # Extensión
                file_size,              # Tamaño
                file_path,              # Ruta en el servidor
                donador,                # Donador
                descripcion,            # Descripción
                consentimiento          # Consentimiento
            ))
            
            # Obtener el ID del documento recién insertado
            documento_id = cursor.lastrowid
            
            # Procesar datos Excel si es un archivo xlsx
            datos_extraidos = []
            if file_extension == 'xlsx':
                try:
                    # Procesar el archivo Excel
                    datos_excel = procesar_excel_plantilla(file_path)
                    metadatos_excel = extraer_metadatos_excel(file_path)
                    
                    # Insertar datos en la tabla datos_exoplanetas
                    if datos_excel:
                        sql_datos = """
                        INSERT INTO datos_exoplanetas 
                        (documento_id, columna_final, origen_tess, origen_kepler, origen_k2, descripcion)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """
                        
                        for dato in datos_excel:
                            cursor.execute(sql_datos, (
                                documento_id,
                                dato['columna_final'],
                                dato.get('origen_tess'),
                                dato.get('origen_kepler'),
                                dato.get('origen_k2'),
                                dato.get('descripcion')
                            ))
                        
                        datos_extraidos = datos_excel
                    
                except Exception as e:
                    logger.error(f"Error procesando Excel {file.filename}: {e}")
                    # No marcamos como error, solo registramos el problema
                    datos_extraidos = [{'error': f'Error procesando datos Excel: {str(e)}'}]
            
            archivos_subidos.append({
                'nombre': file.filename,
                'tipo': file_extension,
                'tamano': file_size,
                'ruta': unique_filename,
                'documento_id': documento_id,
                'datos_extraidos': len(datos_extraidos) if file_extension == 'xlsx' else 0,
                'procesado_excel': file_extension == 'xlsx'
            })
            
            # Agregar datos procesados a la respuesta
            if datos_extraidos:
                datos_procesados.append({
                    'documento_id': documento_id,
                    'nombre_archivo': file.filename,
                    'total_datos': len(datos_extraidos),
                    'datos': datos_extraidos
                })
        
        connection.commit()
        cursor.close()
        connection.close()
        
        response_data = {
            'message': f'Se subieron {len(archivos_subidos)} archivos correctamente',
            'archivos': archivos_subidos,
            'procesados_excel': len([f for f in archivos_subidos if f['procesado_excel']])
        }
        
        if datos_procesados:
            response_data['datos_procesados'] = datos_procesados
        
        if archivos_con_error:
            response_data['errores'] = archivos_con_error
        
        return jsonify(response_data), 200
        
    except Exception as e:
        logger.error(f"Error en upload_files: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
        
@app.route('/api/descargar-todos', methods=['GET'])
def descargar_todos_documentos():
    """Endpoint para descargar todos los documentos en un ZIP"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT id, nombre_archivo, ruta_archivo FROM documentos_exoplanetas")
        
        documentos = cursor.fetchall()
        cursor.close()
        connection.close()
        
        if not documentos:
            return jsonify({'error': 'No hay documentos disponibles'}), 404
        
        # Crear archivo ZIP en memoria
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for documento in documentos:
                if os.path.exists(documento['ruta_archivo']):
                    # Agregar archivo al ZIP
                    zip_file.write(
                        documento['ruta_archivo'], 
                        documento['nombre_archivo']
                    )
        
        zip_buffer.seek(0)
        
        # Enviar ZIP
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"documentos_exoplanetas_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
            mimetype='application/zip'
        )
        
    except Exception as e:
        logger.error(f"Error en descargar_todos_documentos: {e}")
        return jsonify({'error': 'Error al crear el archivo ZIP'}), 500
    
@app.route('/api/total-archivos', methods=['GET'])
def total_archivos():
    """Endpoint simple para obtener solo el total de archivos"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Error de conexión'}), 500
        
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) as total FROM documentos_exoplanetas")
        total = cursor.fetchone()[0]
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'total_archivos': total
        }), 200
        
    except Exception as e:
        logger.error(f"Error en total-archivos: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500

@app.route('/api/total-donadores', methods=['GET'])
def total_donadores():
    """Endpoint simple para obtener solo el total de donadores"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Error de conexión'}), 500
        
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(DISTINCT donador) as total FROM documentos_exoplanetas")
        total = cursor.fetchone()[0]
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'total_donadores': total
        }), 200
        
    except Exception as e:
        logger.error(f"Error en total-donadores: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    
@app.route('/api/descargar-plantilla', methods=['GET'])
def descargar_plantilla():
    """Endpoint para descargar la plantilla Excel"""
    try:
        # Ruta del archivo plantilla
        plantilla_path = os.path.join(os.path.dirname(__file__), 'Plantilla.xlsx')
        
        # Verificar que el archivo existe
        if not os.path.exists(plantilla_path):
            return jsonify({
                'success': False,
                'error': 'Plantilla no disponible'
            }), 404
        
        # Enviar archivo
        return send_file(
            plantilla_path,
            as_attachment=True,
            download_name='Plantilla_Exoplanetas.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error en descargar_plantilla: {e}")
        return jsonify({'error': 'Error al descargar la plantilla'}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)